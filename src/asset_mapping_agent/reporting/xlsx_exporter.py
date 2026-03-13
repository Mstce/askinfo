from __future__ import annotations

import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from asset_mapping_agent.assets import MergedAssetRecord, VerificationResult


KEY_TAGS = {"login_page", "admin_panel", "middleware_console"}
INVALID_STATUSES = {"network_error", "timeout", "refused"}

ASSET_HEADERS = [
    "资产ID",
    "归并键",
    "主机",
    "域名",
    "IP",
    "端口",
    "协议",
    "URL",
    "标题",
    "服务",
    "产品",
    "组织",
    "ICP",
    "国家",
    "省份",
    "城市",
    "相关主机名",
    "标签",
    "来源平台",
    "来源查询",
    "验证方式",
    "验证状态",
    "验证状态码",
    "验证标题",
    "验证详情",
    "验证时间",
    "验证目标",
    "最终目标",
    "是否重点资产",
    "是否新增资产",
    "是否无效资产",
]

ASSET_COLUMN_WIDTHS = {
    "资产ID": 16,
    "归并键": 30,
    "主机": 24,
    "域名": 22,
    "IP": 16,
    "端口": 9,
    "协议": 10,
    "URL": 42,
    "标题": 28,
    "服务": 16,
    "产品": 18,
    "组织": 24,
    "ICP": 20,
    "国家": 10,
    "省份": 10,
    "城市": 10,
    "相关主机名": 30,
    "标签": 26,
    "来源平台": 16,
    "来源查询": 40,
    "验证方式": 14,
    "验证状态": 14,
    "验证状态码": 12,
    "验证标题": 28,
    "验证详情": 30,
    "验证时间": 22,
    "验证目标": 38,
    "最终目标": 38,
    "是否重点资产": 12,
    "是否新增资产": 12,
    "是否无效资产": 12,
}

WRAP_TEXT_HEADERS = {
    "归并键",
    "URL",
    "标题",
    "相关主机名",
    "标签",
    "来源查询",
    "验证标题",
    "验证详情",
    "验证目标",
    "最终目标",
}

CENTER_HEADERS = {
    "端口",
    "协议",
    "国家",
    "省份",
    "城市",
    "验证方式",
    "验证状态",
    "验证状态码",
    "是否重点资产",
    "是否新增资产",
    "是否无效资产",
}

THIN_SIDE = Side(style="thin", color="D9DEE7")
CELL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
BODY_FONT = Font(name="Microsoft YaHei", color="1F2937")
BODY_ALIGNMENT = Alignment(vertical="top")
BODY_ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center")
LINK_FONT = Font(name="Microsoft YaHei", color="0563C1", underline="single")
ZEBRA_FILL = PatternFill("solid", fgColor="F7FBFF")
KEY_ROW_FILL = PatternFill("solid", fgColor="FFF4CE")
NEW_ROW_FILL = PatternFill("solid", fgColor="EAF7EE")
INVALID_ROW_FILL = PatternFill("solid", fgColor="FDEDEC")
SECTION_FILL = PatternFill("solid", fgColor="DCE6F1")
SECTION_FONT = Font(name="Microsoft YaHei", bold=True, color="1F2937")
TITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FONT = Font(name="Microsoft YaHei", bold=True, color="0F172A", size=12)
SUMMARY_TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SUMMARY_SUBTITLE_FILL = PatternFill("solid", fgColor="EAF2F8")
SUMMARY_CARD_LABEL_FILL = PatternFill("solid", fgColor="DCE6F1")
SUMMARY_CARD_VALUE_FILL = PatternFill("solid", fgColor="F8FBFF")
SUMMARY_TITLE_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=14)
SUMMARY_SUBTITLE_FONT = Font(name="Microsoft YaHei", color="334155")
SUMMARY_CARD_LABEL_FONT = Font(name="Microsoft YaHei", bold=True, color="1F2937")
SUMMARY_CARD_VALUE_FONT = Font(name="Microsoft YaHei", bold=True, color="0F172A", size=16)


@dataclass(slots=True)
class AssetWorkbookExportResult:
    output_path: Path
    total_assets: int
    key_assets: int
    new_assets: int
    invalid_assets: int


@dataclass(slots=True)
class AgentWorkbookSummary:
    source_text: str
    subject_name: str
    known_domains: list[str] = field(default_factory=list)
    province: str = ""
    focus: list[str] = field(default_factory=list)
    primary_platforms: list[str] = field(default_factory=list)
    enrichment_platforms: list[str] = field(default_factory=list)
    follow_domain_enrichment: bool = False
    verify_http: bool = True
    verify_tcp: bool = True
    max_results_per_platform: int = 100
    max_primary_platforms: int = 3
    max_enrichment_rounds: int = 2
    max_enrichment_domains_total: int = 10
    max_platform_calls: int = 20
    special_output_format: str = ""
    notes: list[str] = field(default_factory=list)


@dataclass(slots=True)
class AgentWorkbookLogEntry:
    timestamp: str
    stage: str
    message: str
    details: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class AssetWorkbookContext:
    agent_summary: AgentWorkbookSummary | None = None
    execution_logs: list[AgentWorkbookLogEntry] = field(default_factory=list)


@dataclass(slots=True)
class AssetWorkbookExporter:
    summary_sheet_name: str = "汇总"
    agent_summary_sheet_name: str = "Agent摘要"
    execution_log_sheet_name: str = "执行日志摘要"
    final_sheet_name: str = "最终资产清单"
    key_sheet_name: str = "重点资产"
    new_sheet_name: str = "新增资产"
    invalid_sheet_name: str = "无效资产"

    def export(
        self,
        assets: list[MergedAssetRecord],
        output_path: str | Path,
        baseline_keys: Iterable[str] | None = None,
        context: AssetWorkbookContext | None = None,
    ) -> AssetWorkbookExportResult:
        baseline_key_set = {key for key in (baseline_keys or []) if key}
        rows = [self._build_row(asset, baseline_key_set) for asset in assets]
        key_rows = [row for row in rows if row["是否重点资产"] == "是"]
        new_rows = [row for row in rows if row["是否新增资产"] == "是"]
        invalid_rows = [row for row in rows if row["是否无效资产"] == "是"]

        workbook = Workbook()
        workbook.remove(workbook.active)

        self._write_summary_sheet(workbook, rows, key_rows, new_rows, invalid_rows)
        if context and context.agent_summary:
            self._write_agent_summary_sheet(workbook, context.agent_summary)
        if context and context.execution_logs:
            self._write_execution_log_sheet(workbook, context.execution_logs)
        self._write_asset_sheet(workbook, self.final_sheet_name, rows)
        self._write_asset_sheet(workbook, self.key_sheet_name, key_rows)
        self._write_asset_sheet(workbook, self.new_sheet_name, new_rows)
        self._write_asset_sheet(workbook, self.invalid_sheet_name, invalid_rows)

        normalized_path = self._normalize_output_path(output_path)
        normalized_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(normalized_path)

        return AssetWorkbookExportResult(
            output_path=normalized_path,
            total_assets=len(rows),
            key_assets=len(key_rows),
            new_assets=len(new_rows),
            invalid_assets=len(invalid_rows),
        )

    def _build_row(self, asset: MergedAssetRecord, baseline_keys: set[str]) -> dict[str, str | int]:
        latest_verification = self._latest_verification(asset)
        is_key_asset = self._is_key_asset(asset)
        is_new_asset = bool(baseline_keys) and asset.normalized_key not in baseline_keys
        is_invalid_asset = latest_verification.status in INVALID_STATUSES if latest_verification else False

        return {
            "资产ID": asset.asset_id,
            "归并键": asset.normalized_key,
            "主机": asset.host,
            "域名": asset.domain,
            "IP": asset.ip,
            "端口": asset.port or "",
            "协议": asset.scheme,
            "URL": asset.url,
            "标题": asset.title,
            "服务": asset.service,
            "产品": asset.product,
            "组织": asset.org,
            "ICP": asset.icp,
            "国家": asset.geo.country,
            "省份": asset.geo.province,
            "城市": asset.geo.city,
            "相关主机名": " | ".join(asset.hostnames),
            "标签": " | ".join(asset.tags),
            "来源平台": " | ".join(asset.source_platforms),
            "来源查询": " | ".join(asset.source_queries),
            "验证方式": latest_verification.method if latest_verification else "",
            "验证状态": latest_verification.status if latest_verification else "",
            "验证状态码": latest_verification.status_code if latest_verification and latest_verification.status_code is not None else "",
            "验证标题": latest_verification.title if latest_verification else "",
            "验证详情": latest_verification.detail if latest_verification else "",
            "验证时间": latest_verification.verified_at if latest_verification else "",
            "验证目标": latest_verification.url if latest_verification else "",
            "最终目标": latest_verification.final_url if latest_verification else "",
            "是否重点资产": "是" if is_key_asset else "否",
            "是否新增资产": "是" if is_new_asset else "否",
            "是否无效资产": "是" if is_invalid_asset else "否",
        }

    def _write_summary_sheet(
        self,
        workbook: Workbook,
        rows: list[dict[str, str | int]],
        key_rows: list[dict[str, str | int]],
        new_rows: list[dict[str, str | int]],
        invalid_rows: list[dict[str, str | int]],
    ) -> None:
        sheet = workbook.create_sheet(self.summary_sheet_name)
        sheet.sheet_view.showGridLines = False
        metrics = [
            ("总资产数", len(rows)),
            ("重点资产数", len(key_rows)),
            ("新增资产数", len(new_rows)),
            ("无效资产数", len(invalid_rows)),
            ("已验证资产数", sum(1 for row in rows if row["验证状态"])),
            ("HTTP 验证成功数", sum(1 for row in rows if row["验证方式"] == "http_request" and row["验证状态"] == "success")),
            ("TCP 验证成功数", sum(1 for row in rows if row["验证方式"] == "tcp_connect" and row["验证状态"] == "success")),
            ("来源平台数", len(self._count_delimited_values(rows, "来源平台"))),
            ("带标签资产数", sum(1 for row in rows if row["标签"])),
        ]
        platform_counts = self._top_counts(self._count_delimited_values(rows, "来源平台"))
        verification_counts = self._top_counts(self._count_simple_values(rows, "验证状态"))
        tag_counts = self._top_counts(self._count_delimited_values(rows, "标签"))

        for column in "ABCDEF":
            sheet.column_dimensions[column].width = 16

        self._merge_summary_range(sheet, "A1:F1", "资产清单汇总", SUMMARY_TITLE_FILL, SUMMARY_TITLE_FONT)
        self._merge_summary_range(sheet, "A2:F2", "本页展示核心指标与分布概览", SUMMARY_SUBTITLE_FILL, SUMMARY_SUBTITLE_FONT)

        card_slots = [(4, 1), (4, 3), (4, 5), (7, 1), (7, 3), (7, 5), (10, 1), (10, 3), (10, 5)]
        for (row_index, column_index), (metric_name, metric_value) in zip(card_slots, metrics):
            self._write_summary_card(sheet, row_index, column_index, metric_name, metric_value)

        self._write_summary_distribution_block(sheet, 14, 1, "来源平台分布", platform_counts)
        self._write_summary_distribution_block(sheet, 14, 3, "验证状态分布", verification_counts)
        self._write_summary_distribution_block(sheet, 14, 5, "重点标签分布", tag_counts)

        sheet.freeze_panes = "A4"

    def _write_agent_summary_sheet(self, workbook: Workbook, summary: AgentWorkbookSummary) -> None:
        sheet = workbook.create_sheet(self.agent_summary_sheet_name)
        sheet.append(["字段", "值"])
        rows = [
            ("原始任务", summary.source_text),
            ("目标主体", summary.subject_name),
            ("已知域名", " | ".join(summary.known_domains)),
            ("省份", summary.province),
            ("关注重点", " | ".join(summary.focus)),
            ("主查询平台", " | ".join(summary.primary_platforms)),
            ("域名增强平台", " | ".join(summary.enrichment_platforms)),
            ("是否域名补查", self._format_bool(summary.follow_domain_enrichment)),
            ("是否HTTP验证", self._format_bool(summary.verify_http)),
            ("是否TCP验证", self._format_bool(summary.verify_tcp)),
            ("单平台结果上限", summary.max_results_per_platform),
            ("主查询平台上限", summary.max_primary_platforms),
            ("补查轮次上限", summary.max_enrichment_rounds),
            ("补查域名总量上限", summary.max_enrichment_domains_total),
            ("平台调用总预算", summary.max_platform_calls),
            ("输出格式", summary.special_output_format or "xlsx"),
            ("Planner备注", " | ".join(summary.notes)),
        ]
        for key, value in rows:
            sheet.append([key, value])
        self._format_two_column_sheet(sheet)

    def _write_execution_log_sheet(self, workbook: Workbook, logs: list[AgentWorkbookLogEntry]) -> None:
        sheet = workbook.create_sheet(self.execution_log_sheet_name)
        sheet.append(["指标", "值"])
        metrics = [
            ("日志总数", len(logs)),
            ("重试事件数", sum(1 for log in logs if log.stage == "platform_retry")),
            ("降级事件数", sum(1 for log in logs if log.stage == "platform_degraded")),
            ("预算限制事件数", sum(1 for log in logs if log.stage == "budget_limit_reached")),
            ("失败事件数", sum(1 for log in logs if log.stage == "run_failed")),
            ("最后阶段", logs[-1].stage if logs else ""),
        ]
        for key, value in metrics:
            sheet.append([key, value])

        section_start = sheet.max_row + 2
        sheet.append([])
        sheet.append(["序号", "时间", "阶段", "消息", "详情"])
        for index, log in enumerate(logs, start=1):
            sheet.append([index, log.timestamp, log.stage, log.message, self._format_details(log.details)])

        self._format_log_sheet(sheet, header_row=section_start + 1)

    def _write_asset_sheet(self, workbook: Workbook, sheet_name: str, rows: list[dict[str, str | int]]) -> None:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(ASSET_HEADERS)
        for row in rows:
            sheet.append([row.get(header, "") for header in ASSET_HEADERS])
        self._format_asset_sheet(sheet)

    def _append_section(self, sheet, title: str, counts: dict[str, int]) -> None:  # type: ignore[no-untyped-def]
        if not counts:
            return
        sheet.append([])
        header_row = sheet.max_row + 1
        sheet.append([title, "数量"])
        self._style_section_header(sheet, header_row)
        for key, value in counts.items():
            sheet.append([key, value])

    def _format_two_column_sheet(self, sheet) -> None:  # type: ignore[no-untyped-def]
        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 48
        sheet.row_dimensions[1].height = 24
        self._style_header_row(sheet, 1)

        for row_index in range(2, sheet.max_row + 1):
            is_section_header = (
                row_index > 1
                and str(sheet.cell(row=row_index, column=2).value or "").strip() == "数量"
                and bool(str(sheet.cell(row=row_index, column=1).value or "").strip())
            )
            if is_section_header:
                self._style_section_header(sheet, row_index)
                continue
            for cell in sheet[row_index]:
                cell.border = CELL_BORDER
                cell.font = BODY_FONT
                cell.alignment = BODY_ALIGNMENT if cell.column == 2 else BODY_ALIGNMENT_CENTER
            if row_index % 2 == 0:
                for cell in sheet[row_index]:
                    cell.fill = ZEBRA_FILL

    def _merge_summary_range(self, sheet, merged_range: str, value: str, fill, font) -> None:  # type: ignore[no-untyped-def]
        sheet.merge_cells(merged_range)
        start_cell = sheet[merged_range.split(":")[0]]
        start_cell.value = value
        min_col, min_row, max_col, max_row = (
            start_cell.column,
            start_cell.row,
            sheet[merged_range.split(":")[1]].column,
            sheet[merged_range.split(":")[1]].row,
        )
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.fill = fill
                cell.border = CELL_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")
        start_cell.font = font

    def _write_summary_card(self, sheet, start_row: int, start_col: int, title: str, value: int) -> None:  # type: ignore[no-untyped-def]
        sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 1)
        sheet.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=start_row + 1, end_column=start_col + 1)

        title_cell = sheet.cell(row=start_row, column=start_col)
        title_cell.value = title
        title_cell.fill = SUMMARY_CARD_LABEL_FILL
        title_cell.font = SUMMARY_CARD_LABEL_FONT
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        value_cell = sheet.cell(row=start_row + 1, column=start_col)
        value_cell.value = value
        value_cell.fill = SUMMARY_CARD_VALUE_FILL
        value_cell.font = SUMMARY_CARD_VALUE_FONT
        value_cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_index in [start_row, start_row + 1]:
            for col_index in [start_col, start_col + 1]:
                cell = sheet.cell(row=row_index, column=col_index)
                cell.border = CELL_BORDER
                if row_index == start_row:
                    cell.fill = SUMMARY_CARD_LABEL_FILL
                else:
                    cell.fill = SUMMARY_CARD_VALUE_FILL

        sheet.row_dimensions[start_row].height = 22
        sheet.row_dimensions[start_row + 1].height = 28

    def _write_summary_distribution_block(
        self,
        sheet,
        start_row: int,
        start_col: int,
        title: str,
        counts: dict[str, int],
    ) -> None:  # type: ignore[no-untyped-def]
        end_col = start_col + 1
        sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)

        title_cell = sheet.cell(row=start_row, column=start_col)
        title_cell.value = title
        title_cell.fill = SECTION_FILL
        title_cell.font = SECTION_FONT
        title_cell.alignment = HEADER_ALIGNMENT
        for col_index in [start_col, end_col]:
            sheet.cell(row=start_row, column=col_index).border = CELL_BORDER
            sheet.cell(row=start_row, column=col_index).fill = SECTION_FILL

        header_row = start_row + 1
        name_header = sheet.cell(row=header_row, column=start_col)
        count_header = sheet.cell(row=header_row, column=end_col)
        name_header.value = "分类"
        count_header.value = "数量"
        for cell in (name_header, count_header):
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = CELL_BORDER
            cell.alignment = HEADER_ALIGNMENT

        items = list(counts.items()) if counts else [("无数据", 0)]
        for offset, (name, count) in enumerate(items, start=2):
            row_index = start_row + offset
            name_cell = sheet.cell(row=row_index, column=start_col)
            count_cell = sheet.cell(row=row_index, column=end_col)
            name_cell.value = name
            count_cell.value = count
            fill = ZEBRA_FILL if offset % 2 == 0 else PatternFill(fill_type=None)
            for cell in (name_cell, count_cell):
                cell.fill = fill
                cell.border = CELL_BORDER
                cell.font = BODY_FONT
            name_cell.alignment = BODY_ALIGNMENT
            count_cell.alignment = BODY_ALIGNMENT_CENTER

    def _top_counts(self, counts: dict[str, int], limit: int = 8) -> dict[str, int]:
        if len(counts) <= limit:
            return counts
        items = list(counts.items())
        result = dict(items[:limit])
        other_total = sum(value for _, value in items[limit:])
        if other_total:
            result["其他"] = other_total
        return result

    def _format_log_sheet(self, sheet, header_row: int) -> None:  # type: ignore[no-untyped-def]
        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 16
        sheet.column_dimensions["B"].width = 28
        sheet.column_dimensions["C"].width = 22
        sheet.column_dimensions["D"].width = 34
        sheet.column_dimensions["E"].width = 52
        self._style_header_row(sheet, 1)
        self._style_header_row(sheet, header_row)

        for row_index in range(2, sheet.max_row + 1):
            if row_index == header_row:
                continue
            for cell in sheet[row_index]:
                cell.border = CELL_BORDER
                cell.font = BODY_FONT
                cell.alignment = BODY_ALIGNMENT if cell.column != 1 else BODY_ALIGNMENT_CENTER
            if row_index > header_row and (row_index - header_row) % 2 == 1:
                for cell in sheet[row_index]:
                    cell.fill = ZEBRA_FILL

    def _format_asset_sheet(self, sheet) -> None:  # type: ignore[no-untyped-def]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(ASSET_HEADERS))}{max(sheet.max_row, 1)}"
        sheet.row_dimensions[1].height = 26
        self._style_header_row(sheet, 1)

        for column_index, header in enumerate(ASSET_HEADERS, start=1):
            sheet.column_dimensions[get_column_letter(column_index)].width = ASSET_COLUMN_WIDTHS.get(header, 16)

        for row_index in range(2, sheet.max_row + 1):
            row_fill = self._pick_asset_row_fill(sheet, row_index)
            for column_index, header in enumerate(ASSET_HEADERS, start=1):
                cell = sheet.cell(row=row_index, column=column_index)
                cell.border = CELL_BORDER
                cell.font = BODY_FONT
                cell.fill = row_fill or (ZEBRA_FILL if row_index % 2 == 0 else PatternFill(fill_type=None))
                cell.alignment = BODY_ALIGNMENT_CENTER if header in CENTER_HEADERS else BODY_ALIGNMENT
                if header in WRAP_TEXT_HEADERS:
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical="top",
                        wrap_text=True,
                    )
                if header in {"URL", "验证目标", "最终目标"}:
                    self._apply_link_style(cell)
            sheet.row_dimensions[row_index].height = 22

    def _style_header_row(self, sheet, row_index: int) -> None:  # type: ignore[no-untyped-def]
        for cell in sheet[row_index]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = CELL_BORDER
            cell.alignment = HEADER_ALIGNMENT

    def _style_section_header(self, sheet, row_index: int) -> None:  # type: ignore[no-untyped-def]
        for cell in sheet[row_index]:
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            cell.border = CELL_BORDER
            cell.alignment = HEADER_ALIGNMENT

    def _pick_asset_row_fill(self, sheet, row_index: int):  # type: ignore[no-untyped-def]
        is_key = sheet.cell(row=row_index, column=29).value == "是"
        is_new = sheet.cell(row=row_index, column=30).value == "是"
        is_invalid = sheet.cell(row=row_index, column=31).value == "是"
        if is_invalid:
            return INVALID_ROW_FILL
        if is_key:
            return KEY_ROW_FILL
        if is_new:
            return NEW_ROW_FILL
        return None

    def _apply_link_style(self, cell) -> None:  # type: ignore[no-untyped-def]
        value = str(cell.value or "").strip()
        if value.startswith("http://") or value.startswith("https://") or value.startswith("tcp://"):
            cell.hyperlink = value
            cell.font = LINK_FONT

    def _count_delimited_values(self, rows: list[dict[str, str | int]], key: str) -> dict[str, int]:
        counts: dict[str, int] = {}
        for row in rows:
            for item in str(row.get(key, "") or "").split(" | "):
                value = item.strip()
                if not value:
                    continue
                counts[value] = counts.get(value, 0) + 1
        return dict(sorted(counts.items(), key=lambda item: (-item[1], item[0])))

    def _count_simple_values(self, rows: list[dict[str, str | int]], key: str) -> dict[str, int]:
        counts: dict[str, int] = {}
        for row in rows:
            value = str(row.get(key, "") or "").strip()
            if not value:
                continue
            counts[value] = counts.get(value, 0) + 1
        return dict(sorted(counts.items(), key=lambda item: (-item[1], item[0])))

    def _format_details(self, details: dict[str, Any]) -> str:
        if not details:
            return ""
        return json.dumps(details, ensure_ascii=False)

    def _latest_verification(self, asset: MergedAssetRecord) -> VerificationResult | None:
        if not asset.verification_results:
            return None
        return asset.verification_results[-1]

    def _is_key_asset(self, asset: MergedAssetRecord) -> bool:
        for tag in asset.tags:
            if tag in KEY_TAGS or tag.startswith("middleware:"):
                return True
        return False

    def _format_bool(self, value: bool) -> str:
        return "是" if value else "否"

    def _normalize_output_path(self, output_path: str | Path) -> Path:
        path = Path(output_path)
        if not path.suffix:
            return path.with_suffix(".xlsx")
        if path.suffix.lower() == ".xls":
            return path.with_suffix(".xlsx")
        return path
